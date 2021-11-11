from openpyxl import Workbook,load_workbook
from openpyxl.chart import AreaChart,Reference
from openpyxl.utils import get_column_letter
import random

# opens file 
workbook = load_workbook(filename="test.xlsx")

workbook = Workbook()
sheet = workbook.active

rows = [ 
    (None,None,random.randint(1,100),random.randint(1,100)),
    (None,None,random.randint(1,100),random.randint(1,100)),
    (None,None,random.randint(1,100),random.randint(1,100)),
    (None,None,random.randint(1,100),random.randint(1,100)),
    (None,None,random.randint(1,100),random.randint(1,100)),
    (None,None,random.randint(1,100),random.randint(1,100)),
    (None,None,random.randint(1,100),random.randint(1,100)),
]
  
# writes data for the rows above
for row in rows:
    sheet.append(row)

# saves file
workbook.save(filename="test.xlsx")